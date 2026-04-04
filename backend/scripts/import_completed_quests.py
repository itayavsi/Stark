import argparse
import sys
import uuid
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

PROJECT_BACKEND_DIR = Path(__file__).resolve().parents[1]
if str(PROJECT_BACKEND_DIR) not in sys.path:
    sys.path.insert(0, str(PROJECT_BACKEND_DIR))

from services.db import get_connection  # noqa: E402

VALID_STATUSES = {"Open", "Taken", "In Progress", "Done", "Approved", "Stopped", "Cancelled", "ממתין"}
DEFAULT_GROUP = "לווינות"
DEFAULT_FT = "FT1"
DEFAULT_STATUS = "Done"
OPEN_QUESTS_TABLE = "open_quests"
FINISHED_QUESTS_TABLE = "finished_quests"
HEADER_ALIASES = {
    "title": {"title", "quest", "quest_title", "name", "שם", "כותרת"},
    "description": {"description", "details", "notes", "desc", "תיאור", "הערות"},
    "date": {"date", "quest_date", "completed_at", "done_date", "תאריך", "תאריך סיום"},
    "assigned_user": {
        "assigned_user",
        "assigned to",
        "assignee",
        "owner",
        "username",
        "user",
        "מבצע",
        "משויך",
    },
    "group": {"group", "group_name", "team", "צוות", "קבוצה"},
    "year": {"year", "quest_year", "שנה"},
    "ft": {"ft", "ft_name"},
    "shapefile_path": {"shapefile_path", "shape_path", "path", "נתיב"},
}


@dataclass
class ImportRow:
    row_number: int
    quest: dict[str, Any]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Import completed quests from an Excel workbook into the open_quests table."
    )
    parser.add_argument("workbook", help="Path to the .xlsx workbook to import.")
    parser.add_argument("--sheet", help="Worksheet name. Defaults to the active sheet.")
    parser.add_argument("--status", default=DEFAULT_STATUS, help="Status to apply to imported rows.")
    parser.add_argument("--group", default=DEFAULT_GROUP, help="Default group when the sheet does not contain one.")
    parser.add_argument("--ft", default=DEFAULT_FT, help="Default FT when the sheet does not contain one.")
    parser.add_argument(
        "--year",
        type=int,
        help="Default year when the sheet does not contain one and the date cannot be parsed.",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Validate and show what would be inserted without writing to the database.",
    )
    parser.add_argument("--title-column", help="Explicit column header for the quest title.")
    parser.add_argument("--description-column", help="Explicit column header for the quest description.")
    parser.add_argument("--date-column", help="Explicit column header for the quest date.")
    parser.add_argument("--assigned-user-column", help="Explicit column header for the assigned user.")
    parser.add_argument("--group-column", help="Explicit column header for the group.")
    parser.add_argument("--year-column", help="Explicit column header for the year.")
    parser.add_argument("--ft-column", help="Explicit column header for the FT value.")
    parser.add_argument("--shapefile-path-column", help="Explicit column header for the shapefile path.")
    return parser.parse_args()


def _normalize_header(value: Any) -> str:
    return str(value or "").strip().lower().replace("-", "_")


def _build_header_map(header_row: tuple[Any, ...]) -> dict[str, int]:
    header_map: dict[str, int] = {}
    for index, value in enumerate(header_row):
        normalized = _normalize_header(value)
        if normalized:
            header_map[normalized] = index
    return header_map


def _resolve_column(header_map: dict[str, int], explicit: str | None, field_name: str) -> int | None:
    if explicit:
        normalized = _normalize_header(explicit)
        if normalized not in header_map:
            raise ValueError(f"Column '{explicit}' was not found in the workbook headers.")
        return header_map[normalized]

    for alias in HEADER_ALIASES.get(field_name, set()):
        if alias in header_map:
            return header_map[alias]
    return None


def _cell_to_text(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _normalize_date(value: Any) -> str | None:
    if value in (None, ""):
        return None

    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")

    text = str(value).strip()
    if not text:
        return None

    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(text, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue

    if "T" in text:
        return text.split("T", 1)[0]

    return text


def _normalize_year(value: Any, date_text: str | None, default_year: int | None) -> int:
    if value not in (None, ""):
        try:
            return int(str(value).strip())
        except ValueError as exc:
            raise ValueError(f"Invalid year value '{value}'.") from exc

    if date_text:
        try:
            return int(date_text[:4])
        except ValueError:
            pass

    if default_year is not None:
        return default_year

    return datetime.now().year


def _row_value(values: tuple[Any, ...], index: int | None) -> Any:
    if index is None or index >= len(values):
        return None
    return values[index]


def _load_rows(args: argparse.Namespace) -> list[ImportRow]:
    workbook_path = Path(args.workbook).expanduser().resolve()
    if workbook_path.suffix.lower() != ".xlsx":
        raise ValueError("Only .xlsx workbooks are supported right now.")
    if not workbook_path.exists():
        raise FileNotFoundError(f"Workbook not found: {workbook_path}")

    workbook = load_workbook(workbook_path, data_only=True)
    sheet = workbook[args.sheet] if args.sheet else workbook.active

    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise ValueError("The worksheet is empty.")

    header_map = _build_header_map(rows[0])
    title_index = _resolve_column(header_map, args.title_column, "title")
    if title_index is None:
        raise ValueError("Could not detect a title column. Use --title-column to specify it.")

    column_indexes = {
        "title": title_index,
        "description": _resolve_column(header_map, args.description_column, "description"),
        "date": _resolve_column(header_map, args.date_column, "date"),
        "assigned_user": _resolve_column(header_map, args.assigned_user_column, "assigned_user"),
        "group": _resolve_column(header_map, args.group_column, "group"),
        "year": _resolve_column(header_map, args.year_column, "year"),
        "ft": _resolve_column(header_map, args.ft_column, "ft"),
        "shapefile_path": _resolve_column(header_map, args.shapefile_path_column, "shapefile_path"),
    }

    imported_rows: list[ImportRow] = []
    for sheet_row_number, values in enumerate(rows[1:], start=2):
        raw_title = _cell_to_text(_row_value(values, column_indexes["title"]))
        if not raw_title:
            continue

        date_text = _normalize_date(_row_value(values, column_indexes["date"])) or datetime.now().strftime("%Y-%m-%d")
        quest = {
            "id": str(uuid.uuid4()),
            "title": raw_title,
            "description": _cell_to_text(_row_value(values, column_indexes["description"])) or "",
            "status": args.status,
            "priority": "רגיל",
            "date": date_text,
            "assigned_user": _cell_to_text(_row_value(values, column_indexes["assigned_user"])),
            "shapefile_path": _cell_to_text(_row_value(values, column_indexes["shapefile_path"])),
            "group": _cell_to_text(_row_value(values, column_indexes["group"])) or args.group,
            "year": _normalize_year(_row_value(values, column_indexes["year"]), date_text, args.year),
            "ft": _cell_to_text(_row_value(values, column_indexes["ft"])) or args.ft,
        }
        imported_rows.append(ImportRow(row_number=sheet_row_number, quest=quest))

    return imported_rows


def _quest_exists(cur: Any, quest: dict[str, Any]) -> bool:
    params = {
        "title": quest["title"],
        "date": quest["date"],
        "assigned_user": quest["assigned_user"],
        "group_name": quest["group"],
        "year": quest["year"],
        "ft": quest["ft"],
    }
    for table_name in (OPEN_QUESTS_TABLE, FINISHED_QUESTS_TABLE):
        cur.execute(
            f"""
            SELECT 1
            FROM {table_name}
            WHERE title = %(title)s
              AND date = %(date)s
              AND COALESCE(assigned_user, '') = COALESCE(%(assigned_user)s, '')
              AND group_name = %(group_name)s
              AND year = %(year)s
              AND ft = %(ft)s
            LIMIT 1;
            """,
            params,
        )
        if cur.fetchone() is not None:
            return True
    return False


def _insert_quest(cur: Any, quest: dict[str, Any]) -> None:
    target_table = FINISHED_QUESTS_TABLE if quest.get("status") in {"Done", "Approved"} else OPEN_QUESTS_TABLE
    cur.execute(
        f"""
        INSERT INTO {target_table} (
            id, title, description, status, "תעדוף", date, assigned_user,
            shapefile_path, group_name, year, ft,
            geometry_status, geometry_source_path, geometry_feature_count, geometry_updated_at
        )
        VALUES (
            %(id)s, %(title)s, %(description)s, %(status)s, %(priority)s, %(date)s, %(assigned_user)s,
            %(shapefile_path)s, %(group_name)s, %(year)s, %(ft)s,
            %(geometry_status)s, %(geometry_source_path)s, %(geometry_feature_count)s, NOW()
        );
        """,
        {
            "id": quest["id"],
            "title": quest["title"],
            "description": quest["description"],
            "status": quest["status"],
            "priority": quest.get("priority", "רגיל"),
            "date": quest["date"],
            "assigned_user": quest["assigned_user"],
            "shapefile_path": quest["shapefile_path"],
            "group_name": quest["group"],
            "year": quest["year"],
            "ft": quest["ft"],
            "geometry_status": "pending" if quest.get("shapefile_path") else "missing",
            "geometry_source_path": quest.get("shapefile_path"),
            "geometry_feature_count": 0,
        },
    )


def main() -> int:
    args = parse_args()
    if args.status not in VALID_STATUSES:
        raise ValueError(f"Invalid status '{args.status}'. Expected one of: {', '.join(sorted(VALID_STATUSES))}.")

    import_rows = _load_rows(args)
    if not import_rows:
        print("No importable rows were found in the workbook.")
        return 0

    inserted = 0
    skipped = 0

    with get_connection() as conn:
        with conn.cursor() as cur:
            for import_row in import_rows:
                if _quest_exists(cur, import_row.quest):
                    skipped += 1
                    print(f"Skipping row {import_row.row_number}: quest already exists.")
                    continue

                if args.dry_run:
                    inserted += 1
                    print(f"Would import row {import_row.row_number}: {import_row.quest['title']}")
                    continue

                _insert_quest(cur, import_row.quest)
                inserted += 1
                print(f"Imported row {import_row.row_number}: {import_row.quest['title']}")

        if args.dry_run:
            conn.rollback()
        else:
            conn.commit()

    mode = "Dry run complete" if args.dry_run else "Import complete"
    print(f"{mode}. {inserted} rows ready/inserted, {skipped} duplicates skipped.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
